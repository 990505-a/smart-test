"""Excel processor using openpyxl (ExcelProcessor).

Per D-07: Converts Excel (.xlsx/.xls) sheets to Markdown tables.
Uses openpyxl with data_only=True to read cell values (not formulas).
Each sheet is rendered as a separate Markdown table with a header.
"""

from __future__ import annotations

import logging
import tempfile
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excel processor that converts spreadsheet data to Markdown tables.

    Reads .xlsx files using openpyxl with data_only=True (returns calculated
    values instead of formulas). Each sheet is converted to a Markdown table
    prefixed with a sheet heading.
    """

    def extract_text(self, excel_data: bytes, filename: str = "data.xlsx") -> str:
        """Extract Markdown table representation from Excel bytes.

        Args:
            excel_data: Raw Excel file bytes.
            filename: Original filename for logging purposes.

        Returns:
            Markdown string with all sheets as tables, or error message.
        """
        if not excel_data:
            return ""

        temp_path = None
        try:
            # Write to temp file (openpyxl needs a file path)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(excel_data)
                temp_path = tmp.name

            wb = openpyxl.load_workbook(temp_path, data_only=True)
            parts: list[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                # Build Markdown table
                parts.append(f"### Sheet: {sheet_name}")

                # Header row
                header = rows[0]
                header_cells = [str(cell) if cell is not None else "" for cell in header]
                parts.append("| " + " | ".join(header_cells) + " |")
                parts.append("| " + " | ".join(["---"] * len(header_cells)) + " |")

                # Data rows
                for row in rows[1:]:
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    # Pad row to match header column count
                    while len(cells) < len(header_cells):
                        cells.append("")
                    parts.append("| " + " | ".join(cells[:len(header_cells)]) + " |")

                parts.append("")  # Blank line between sheets

            wb.close()
            return "\n".join(parts).strip()

        except Exception as e:
            logger.error("ExcelProcessor: failed to process %s: %s", filename, e)
            return f"Excel processing error: {e}"
        finally:
            if temp_path is not None:
                try:
                    import os
                    os.unlink(temp_path)
                except Exception:
                    pass
