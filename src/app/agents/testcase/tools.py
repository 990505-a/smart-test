"""TestCase Agent tools - Excel export with professional formatting.

Per D-09: LLM generates Markdown test cases -> backend parses Markdown -> openpyxl writes Excel.
Per D-10: TC numbering convention TC-[PROJECT]-[MODULE]-[NNN].
Per D-11: Professional formatting with header style, borders, alignment, auto-wrap.
Per D-12: Unified export function supporting excel/csv/json/markdown formats.
Per D-13: CSV export with UTF-8 BOM for ZenTao/TestRail compatibility.
Per D-14: JSON export in Jira Xray format.
"""
import csv
import io
import json
from pathlib import Path
from typing import Any

from langchain.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- Style constants (D-11) ---
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column widths (D-11)
_DEFAULT_COLUMN_WIDTHS = {
    "A": 18, "B": 35, "C": 14, "D": 12, "E": 10,
    "F": 30, "G": 40, "H": 30, "I": 40, "J": 20,
}

# 10 standard columns
HEADERS = [
    "用例编号", "用例标题", "所属模块", "用例类型", "优先级",
    "前置条件", "测试步骤", "测试数据", "预期结果", "备注",
]


# --- Field extraction helpers (EXPT-04) ---

def _extract_field(case: dict, *keys: str, default: Any = "") -> Any:
    """Extract a value from a dict using multiple candidate keys.

    Iterates over the provided keys and returns the value of the first match.
    If none match, returns the default.
    """
    for key in keys:
        if key in case:
            return case[key]
    return default


def _flatten_steps(steps: list[dict] | None) -> str:
    """Format a list of step dicts as numbered text.

    Each step is formatted as: "N. action [target]" with optional data.
    """
    if not steps:
        return ""
    lines = []
    for step in steps:
        seq = step.get("seq", step.get("step", len(lines) + 1))
        action = step.get("action", step.get("操作描述", ""))
        target = step.get("target", step.get("操作对象", ""))
        data = step.get("data", "")
        line = f"{seq}. {action}"
        if target:
            line += f" [{target}]"
        if data:
            line += f"（数据：{data}）"
        lines.append(line)
    return "\n".join(lines)


def _flatten_test_data(test_data: dict | str | None) -> str:
    """Format test data as 'key: value' lines or return string directly."""
    if not test_data:
        return ""
    if isinstance(test_data, str):
        return test_data
    lines = [f"{k}: {v}" for k, v in test_data.items()]
    return "\n".join(lines)


def _flatten_expected_results(expected_results: list[str] | str | None) -> str:
    """Format expected results as numbered list or return string directly."""
    if not expected_results:
        return ""
    if isinstance(expected_results, str):
        return expected_results
    lines = []
    for idx, result in enumerate(expected_results, start=1):
        lines.append(f"{idx}. {result}")
    return "\n".join(lines)


def _flatten_preconditions(preconditions: list[str] | str | None) -> str:
    """Format preconditions as numbered list or return string directly."""
    if not preconditions:
        return ""
    if isinstance(preconditions, str):
        return preconditions
    lines = []
    for idx, cond in enumerate(preconditions, start=1):
        lines.append(f"{idx}. {cond}")
    return "\n".join(lines)


def _export_csv(test_cases: list[dict], output_path: str) -> str:
    """Export test cases as CSV with UTF-8 BOM for ZenTao/TestRail compatibility.

    Per D-13: UTF-8 BOM encoding, comma delimited, double-quote escaping,
    10 standard columns compatible with ZenTao and TestRail import.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow(HEADERS)

    for case in test_cases:
        writer.writerow([
            _extract_field(case, "id", "用例编号"),
            _extract_field(case, "title", "用例标题"),
            _extract_field(case, "module", "所属模块"),
            _extract_field(case, "type", "用例类型"),
            _extract_field(case, "priority", "优先级"),
            _flatten_preconditions(_extract_field(case, "preconditions", "前置条件", default=None)),
            _flatten_steps(_extract_field(case, "steps", "测试步骤", default=None)),
            _flatten_test_data(_extract_field(case, "test_data", "测试数据", default=None)),
            _flatten_expected_results(_extract_field(case, "expected_results", "预期结果", default=None)),
            _extract_field(case, "remarks", "备注"),
        ])

    with open(output_path, "wb") as f:
        f.write(b'\xef\xbb\xbf')  # UTF-8 BOM per D-13
        f.write(buf.getvalue().encode("utf-8"))

    return str(output_path.resolve())


def _export_json(test_cases: list[dict], output_path: str) -> str:
    """Export test cases as JSON compatible with Jira Xray.

    Per D-14: Format is {"testCases": [{"testCaseKey": ..., "summary": ..., "steps": [...]}]}.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    xray_cases = []
    for case in test_cases:
        steps = _extract_field(case, "steps", "测试步骤", default=None) or []
        xray_steps = []
        if isinstance(steps, list):
            for step in steps:
                xray_steps.append({
                    "action": step.get("action", step.get("操作描述", "")),
                    "result": step.get("expected", step.get("预期结果", "")),
                    "data": step.get("data", ""),
                })

        xray_cases.append({
            "testCaseKey": _extract_field(case, "id", "用例编号"),
            "summary": _extract_field(case, "title", "用例标题"),
            "type": _extract_field(case, "type", "用例类型"),
            "priority": _extract_field(case, "priority", "优先级"),
            "status": "DRAFT",
            "folder": _extract_field(case, "module", "所属模块"),
            "steps": xray_steps,
            "preconditions": _flatten_preconditions(
                _extract_field(case, "preconditions", "前置条件", default=None)
            ),
            "labels": [_extract_field(case, "module", "所属模块")],
        })

    data = {"testCases": xray_cases}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return str(output_path.resolve())


def _export_markdown(test_cases: list[dict], output_path: str) -> str:
    """Export test cases as Markdown table format."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = ["# 测试用例\n"]
    lines.append("| " + " | ".join(HEADERS) + " |")
    lines.append("| " + " | ".join("---" for _ in HEADERS) + " |")

    for case in test_cases:
        row = [
            _extract_field(case, "id", "用例编号"),
            _extract_field(case, "title", "用例标题"),
            _extract_field(case, "module", "所属模块"),
            _extract_field(case, "type", "用例类型"),
            _extract_field(case, "priority", "优先级"),
            _flatten_preconditions(_extract_field(case, "preconditions", "前置条件", default=None)).replace("\n", " "),
            _flatten_steps(_extract_field(case, "steps", "测试步骤", default=None)).replace("\n", " "),
            _flatten_test_data(_extract_field(case, "test_data", "测试数据", default=None)).replace("\n", " "),
            _flatten_expected_results(_extract_field(case, "expected_results", "预期结果", default=None)).replace("\n", " "),
            _extract_field(case, "remarks", "备注"),
        ]
        lines.append("| " + " | ".join(str(v).replace("|", "\\|") for v in row) + " |")

    content = "\n".join(lines)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    return str(output_path.resolve())


@tool
def export_test_cases_to_excel(
    test_cases: list[dict[str, Any]],
    output_path: str,
    sheet_name: str = "测试用例",
) -> str:
    """Export test cases to a professionally formatted Excel file.

    Supports field mapping from multiple key names (EN/CN) per EXPT-04.
    Applies professional styles per D-11.

    Args:
        test_cases: List of test case dicts. Each dict may use EN keys (id, title,
            module, type, priority, preconditions, steps, test_data,
            expected_results, remarks) or CN keys (用例编号, 用例标题, 所属模块,
            用例类型, 优先级, 前置条件, 测试步骤, 测试数据, 预期结果, 备注).
        output_path: File path for the output .xlsx file.
        sheet_name: Name of the Excel sheet. Defaults to "测试用例".

    Returns:
        Absolute path of the created Excel file.

    Raises:
        ValueError: If test_cases list is empty.
    """
    if not test_cases:
        raise ValueError("测试用例列表为空，无法导出 Excel。")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表。")
    ws.title = sheet_name

    # Append header row
    ws.append(HEADERS)

    # Apply header styles
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGNMENT_CENTER
        cell.border = _BORDER

    # Write data rows
    for case in test_cases:
        row = [
            _extract_field(case, "id", "用例编号"),
            _extract_field(case, "title", "用例标题"),
            _extract_field(case, "module", "所属模块"),
            _extract_field(case, "type", "用例类型"),
            _extract_field(case, "priority", "优先级"),
            _flatten_preconditions(_extract_field(case, "preconditions", "前置条件", default=None)),
            _flatten_steps(_extract_field(case, "steps", "测试步骤", default=None)),
            _flatten_test_data(_extract_field(case, "test_data", "测试数据", default=None)),
            _flatten_expected_results(_extract_field(case, "expected_results", "预期结果", default=None)),
            _extract_field(case, "remarks", "备注"),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER

    # Set column widths
    for col_letter, width in _DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Set row heights
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    wb.save(str(output_path))
    return str(output_path.resolve())


@tool
def export_test_cases(
    test_cases: list[dict[str, Any]],
    output_path: str,
    format: str = "excel",
    sheet_name: str = "测试用例",
) -> str:
    """Export test cases in multiple formats.

    Per D-12: Unified export function replacing separate format-specific functions.
    Supported formats: excel (default), csv, json, markdown.

    Args:
        test_cases: List of test case dicts with EN or CN field names.
        output_path: File path for the output file.
        format: Output format - "excel", "csv", "json", or "markdown".
        sheet_name: Excel sheet name (only used for excel format).

    Returns:
        Absolute path of the created file.

    Raises:
        ValueError: If test_cases list is empty or format is unsupported.
    """
    if not test_cases:
        raise ValueError("测试用例列表为空，无法导出。")

    fmt = format.lower().strip()
    if fmt == "excel":
        return export_test_cases_to_excel(test_cases, output_path, sheet_name)
    elif fmt == "csv":
        return _export_csv(test_cases, output_path)
    elif fmt == "json":
        return _export_json(test_cases, output_path)
    elif fmt == "markdown":
        return _export_markdown(test_cases, output_path)
    else:
        raise ValueError(f"不支持的导出格式: {format}。支持: excel, csv, json, markdown")
