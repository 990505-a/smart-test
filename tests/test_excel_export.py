"""Tests for Excel export tool and field extraction."""
import pytest
import tempfile
from pathlib import Path

# Import the tool and helpers
# export_test_cases_to_excel is a LangChain StructuredTool (@tool decorator).
# Use .func to access the underlying Python function for direct testing.
from src.app.agents.testcase.tools import (
    export_test_cases_to_excel as _excel_tool,
    _extract_field,
    _flatten_steps,
    _flatten_test_data,
    _flatten_expected_results,
    _flatten_preconditions,
)

# Get the raw function from the StructuredTool for direct invocation
export_test_cases_to_excel = _excel_tool.func

# --- Field extraction tests (EXPT-04) ---


def test_extract_field_english_key():
    assert _extract_field({"title": "Login test"}, "title", "用例标题") == "Login test"


def test_extract_field_chinese_key():
    assert _extract_field({"用例标题": "登录测试"}, "title", "用例标题") == "登录测试"


def test_extract_field_default():
    assert _extract_field({}, "title", "用例标题") == ""
    assert _extract_field({}, "title", "用例标题", default="N/A") == "N/A"


def test_flatten_steps():
    steps = [{"seq": 1, "action": "Open page", "target": "URL"}, {"seq": 2, "action": "Click button"}]
    result = _flatten_steps(steps)
    assert "1. Open page [URL]" in result
    assert "2. Click button" in result


def test_flatten_steps_none():
    assert _flatten_steps(None) == ""


def test_flatten_test_data_dict():
    result = _flatten_test_data({"username": "admin", "password": "123"})
    assert "username: admin" in result
    assert "password: 123" in result


def test_flatten_test_data_string():
    assert _flatten_test_data("raw data") == "raw data"


def test_flatten_expected_results():
    result = _flatten_expected_results(["Page loads", "Status 200"])
    assert "1. Page loads" in result
    assert "2. Status 200" in result


def test_flatten_preconditions():
    result = _flatten_preconditions(["User exists", "DB ready"])
    assert "1. User exists" in result
    assert "2. DB ready" in result


# --- Excel export tests (EXPT-01, EXPT-02) ---


def test_export_creates_excel_file():
    test_cases = [
        {
            "id": "TC-CRM-LOGIN-001",
            "title": "Valid login",
            "module": "Authentication",
            "type": "功能测试",
            "priority": "P0",
            "preconditions": ["User registered"],
            "steps": [{"seq": 1, "action": "Enter credentials"}],
            "test_data": {"username": "admin"},
            "expected_results": ["Redirect to dashboard"],
            "remarks": "REQ-001",
        }
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "test_output.xlsx"
        result = export_test_cases_to_excel(test_cases, str(output))
        assert Path(result).exists()
        assert result.endswith(".xlsx")


def test_export_excel_has_headers():
    test_cases = [{"id": "TC-001", "title": "Test"}]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "test.xlsx"
        export_test_cases_to_excel(test_cases, str(output))
        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws.title == "测试用例"
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert headers[0] == "用例编号"
        assert headers[1] == "用例标题"


def test_export_excel_header_styles():
    test_cases = [{"id": "TC-001", "title": "Test"}]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "test.xlsx"
        export_test_cases_to_excel(test_cases, str(output))
        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.fill.start_color.rgb == "00366092"  # PatternFill format
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "00FFFFFF"


def test_export_empty_cases_raises():
    with tempfile.TemporaryDirectory() as tmpdir:
        with pytest.raises(ValueError, match="测试用例列表为空"):
            export_test_cases_to_excel([], str(Path(tmpdir) / "test.xlsx"))


def test_tc_numbering_convention():
    """EXPT-02: TC numbering follows TC-[PROJECT]-[MODULE]-[NNN] pattern."""
    test_cases = [
        {"id": "TC-CRM-LOGIN-001", "title": "Test 1"},
        {"id": "TC-OMS-ORDER-012", "title": "Test 2"},
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "test.xlsx"
        export_test_cases_to_excel(test_cases, str(output))
        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "TC-CRM-LOGIN-001"
        assert ws.cell(row=3, column=1).value == "TC-OMS-ORDER-012"


def test_field_extraction_nested_formats():
    """EXPT-04: Field extraction handles both EN and CN key names."""
    test_cases = [
        {"用例编号": "TC-001", "用例标题": "CN key test", "所属模块": "Auth"},
        {"id": "TC-002", "title": "EN key test", "module": "Login"},
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "test.xlsx"
        export_test_cases_to_excel(test_cases, str(output))
        from openpyxl import load_workbook
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=3, column=1).value == "TC-002"
